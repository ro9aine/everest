from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


class ExcelInputError(Exception):
    """Base exception for Excel input validation errors."""


class ExcelFileNotFoundError(ExcelInputError):
    """Raised when the requested workbook path does not exist."""


class ExcelSheetNotFoundError(ExcelInputError):
    """Raised when the requested worksheet cannot be found."""


class ExcelColumnNotFoundError(ExcelInputError):
    """Raised when the requested header is missing from the worksheet."""


@dataclass(slots=True, frozen=True)
class InputRow:
    row_number: int
    value: str


@dataclass(slots=True, frozen=True)
class InputBatch:
    rows: tuple[InputRow, ...]
    skipped_empty: int


class ExcelInputReader:
    """Reads identifiers from XLSX files using the first row as the header row."""

    def read(self, path: str, *, sheet: str | None = None, column: str) -> InputBatch:
        workbook_path = Path(path)
        if not workbook_path.exists():
            raise ExcelFileNotFoundError(f"Input file does not exist: {workbook_path}")
        if workbook_path.suffix.lower() != ".xlsx":
            raise ExcelInputError("Input file must be an .xlsx workbook")

        try:
            workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        except Exception as exc:
            raise ExcelInputError(f"Failed to open workbook: {workbook_path}") from exc

        try:
            worksheet = self._select_sheet(workbook, sheet)
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if header_row is None:
                return InputBatch(rows=(), skipped_empty=0)

            column_index = self._find_column_index(header_row, column)
            rows: list[InputRow] = []
            skipped_empty = 0
            for row_number, row in enumerate(
                worksheet.iter_rows(min_row=2, values_only=True),
                start=2,
            ):
                cell_value = row[column_index] if column_index < len(row) else None
                normalized = self._normalize_cell_value(cell_value)
                if normalized is None:
                    skipped_empty += 1
                    continue
                rows.append(InputRow(row_number=row_number, value=normalized))

            return InputBatch(rows=tuple(rows), skipped_empty=skipped_empty)
        finally:
            workbook.close()

    @staticmethod
    def _select_sheet(workbook, sheet: str | None):
        if sheet is None:
            return workbook[workbook.sheetnames[0]]
        if sheet not in workbook.sheetnames:
            raise ExcelSheetNotFoundError(f"Worksheet '{sheet}' was not found in the workbook")
        return workbook[sheet]

    @staticmethod
    def _find_column_index(header_row: tuple[object, ...], column: str) -> int:
        normalized_column = " ".join(str(column).split()).casefold()
        for index, value in enumerate(header_row):
            current = " ".join(str(value or "").split()).casefold()
            if current == normalized_column:
                return index
        raise ExcelColumnNotFoundError(f"Column '{column}' was not found in the worksheet header")

    @staticmethod
    def _normalize_cell_value(value: object) -> str | None:
        if value is None:
            return None
        normalized = " ".join(str(value).split()).strip()
        return normalized or None
