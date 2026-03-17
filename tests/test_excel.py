from __future__ import annotations

from openpyxl import Workbook

import pytest

from app.excel import ExcelColumnNotFoundError, ExcelInputReader, ExcelSheetNotFoundError


def test_excel_reader_reads_non_empty_rows(tmp_path) -> None:
    path = tmp_path / "input.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(["INN"])
    worksheet.append([" 1234567890 "])
    worksheet.append([None])
    worksheet.append([""])
    worksheet.append(["0987654321"])
    workbook.save(path)

    result = ExcelInputReader().read(str(path), sheet="Sheet1", column="INN")

    assert result.skipped_empty == 2
    assert [row.value for row in result.rows] == ["1234567890", "0987654321"]
    assert [row.row_number for row in result.rows] == [2, 5]


def test_excel_reader_raises_for_missing_column(tmp_path) -> None:
    path = tmp_path / "input.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["CASE"])
    workbook.save(path)

    with pytest.raises(ExcelColumnNotFoundError):
        ExcelInputReader().read(str(path), column="INN")


def test_excel_reader_raises_for_missing_sheet(tmp_path) -> None:
    path = tmp_path / "input.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Main"
    worksheet.append(["INN"])
    workbook.save(path)

    with pytest.raises(ExcelSheetNotFoundError):
        ExcelInputReader().read(str(path), sheet="Sheet1", column="INN")
